import openpyxl
import re
import time

def modify_eyemark_budget(eyemark_budget_path, amended_budget_path, modified_eyemark_path):
    # Load the workbooks
    eyemark_wb = openpyxl.load_workbook(eyemark_budget_path)
    amended_wb = openpyxl.load_workbook(amended_budget_path)

    # Access the active sheets
    eyemark_sheet = eyemark_wb.active
    amended_sheet = amended_wb.active

    # Regular expression pattern for matching entries starting with 'ERGP' followed by numbers
    pattern = re.compile(r'^ERGP\d+$')

    # Create a set for storing matching entries from the Amended 2020 Budget
    amended_entries = set()

    # Collect matching entries from the Amended 2020 Budget, specifically from column A, up to row 74223
    print("Collecting ERGP Codes from amended budget...")
    x_start = time.time()
    for row in amended_sheet.iter_rows(min_row=1, max_row=74223, min_col=1, max_col=1, values_only=True):
        cell_value = row[0]
        # Convert cell value to string before applying regex
        if cell_value is not None and pattern.match(str(cell_value)):
            amended_entries.add(str(cell_value))

    x_end = time.time()
    print(f"Finished collecting ERGP codes in: {x_end - x_start} seconds.")
    print("    ")

    # Loop through the specified range in the Eyemark Budget Template and modify if needed
    print("Validating ERGP codes in Eyemark budget...")
    y_start = time.time()
    for row in range(2, 9090):
        cell_value = eyemark_sheet.cell(row=row, column=1).value
        if cell_value and str(cell_value) not in amended_entries:
            for col in range(1, 4):
                eyemark_sheet.cell(row=row, column=col, value='DNF')

    y_end = time.time()
    print(f"Finished validating ERGP codes in: {y_end - y_start}")

    # Save the modified workbook
    eyemark_wb.save(modified_eyemark_path)

# File paths (update these paths to the location of your files)
eyemark_budget_path = 'eyemark_budget.xlsx'
amended_budget_path = 'amended_budget.xlsx'
modified_eyemark_path = 'modified_eyemark_budget.xlsx'

# Run the function
modify_eyemark_budget(eyemark_budget_path, amended_budget_path, modified_eyemark_path)

